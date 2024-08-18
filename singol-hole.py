import os
import logging
from typing import List, Tuple
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side
import pandas as pd
import math
import argparse
import tkinter as tk
from tkinter import filedialog
from tkinter import filedialog, messagebox
import traceback
import sys
from openpyxl.styles import numbers

# Configuration
INPUT_FILE = ''
OUTPUT_FILE = ''
LOGO_FILE = '萬頂圖樣.png'
MAIN_SHEET = '工作表1'
PROJECT_NAME_CELL = (0,1)

# Constants
THIN_BORDER = Side(border_style='thin')
MEDIUM_BORDER = Side(border_style='medium')

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')



# 配置和常量部分保持不變

# 設置日誌
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def print_error_with_line(e):
    exc_type, exc_obj, exc_tb = sys.exc_info()
    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
    print(f"錯誤類型: {exc_type.__name__}")
    print(f"文件名: {fname}")
    print(f"行號: {exc_tb.tb_lineno}")
    print(f"錯誤訊息: {str(e)}")
    print("\n完整的錯誤追蹤:")
    traceback.print_exc()

class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel 處理程序")
        self.geometry("300x100")
        self.create_widgets()

    def create_widgets(self):
        self.start_button = tk.Button(self, text="單孔柱狀圖報告建立", command=self.start_processing)
        self.start_button.pack(expand=True)

    def start_processing(self):
        INPUT_FILE, OUTPUT_FILE = self.select_files()
        if not INPUT_FILE or not OUTPUT_FILE:
            messagebox.showinfo("信息", "文件選擇已取消")
            return

        self.process_files(INPUT_FILE, OUTPUT_FILE)

    def select_files(self):
        INPUT_FILE = filedialog.askopenfilename(title="選擇輸入Excel文件", filetypes=[("Excel files", "*.xlsx")])
        if not INPUT_FILE:
            return None, None

        OUTPUT_FILE = filedialog.asksaveasfilename(title="保存輸出Excel文件", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not OUTPUT_FILE:
            return None, None

        return INPUT_FILE, OUTPUT_FILE
    def print_error_with_line(e):
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print(f"錯誤類型: {exc_type.__name__}")
        print(f"文件名: {fname}")
        print(f"行號: {exc_tb.tb_lineno}")
        print(f"錯誤訊息: {str(e)}")
        print("\n完整的錯誤追蹤:")
        traceback.print_exc()


    def process_files(self, INPUT_FILE, OUTPUT_FILE):
        try:
            xl, sheet_names, project_name = load_excel_file(INPUT_FILE)
            new_wb = create_new_workbook(sheet_names)

            for sheet_name in sheet_names:
                process_worksheet(sheet_name, xl, new_wb, project_name)

            if os.path.exists(OUTPUT_FILE):
                os.remove(OUTPUT_FILE)
            
            new_wb.save(OUTPUT_FILE)
            messagebox.showinfo("成功", f"成功創建 {OUTPUT_FILE}")
        except Exception as e:
            error_msg = f"發生錯誤: {e}"
            print_error_with_line(e)  # 在終端機輸出詳細錯誤信息
            logging.error(error_msg)  # 記錄到日誌
            messagebox.showerror("錯誤", error_msg)  # GUI顯示

def parse_arguments():
    parser = argparse.ArgumentParser(description='Process Excel file for geological data.')
    parser.add_argument('--input', default=INPUT_FILE, help='Input Excel file name')
    parser.add_argument('--output', default=OUTPUT_FILE, help='Output Excel file name')
    return parser.parse_args()

def load_excel_file(filename: str) -> Tuple[pd.ExcelFile, List[str], str]:
    try:
        xl = pd.ExcelFile(filename)
        ws = pd.read_excel(xl, MAIN_SHEET, header=None)
        project_name = ws.iloc[PROJECT_NAME_CELL[0], PROJECT_NAME_CELL[1]]  # Access the cell for project name
        print(project_name)
        sheet_names = xl.sheet_names
        sheet_names.remove(MAIN_SHEET)
        return xl, sheet_names, project_name
    except Exception as e:
        logging.error(f"Error loading Excel file: {e}")
        raise


def create_new_workbook(sheet_names: List[str]) -> Workbook:
    new_wb = Workbook()
    new_wb.remove(new_wb.active)
    for sheet_name in sheet_names:
        new_wb.create_sheet(title=sheet_name)
    return new_wb

def adjust_column_width(ws):
    column_widths = {
        'A': 6.11, 'B': 1, 'C': 6, 'D': 8, 'E': 6, 'F': 6, 'G': 6, 'H': 6, 'I': 35,
        'J': 12, 'K': 8, 'L': 8, 'M': 8, 'N': 8, 'O': 9, 'P': 9, 'Q': 9, 'R': 9,
        'S': 9, 'T': 9, 'U': 10, 'V': 9, 'W': 9, 'X': 12
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

def merge_cells(ws, start_row: int):
    merge_cells_instructions = [
        (start_row, 1, start_row+2, 24),
        (start_row+3, 1, start_row+3, 24),
        (start_row+4, 1, start_row+4, 24),
        (start_row+9, 5, start_row+9, 8),
        (start_row+10, 5, start_row+10, 8),
        (start_row+11, 5, start_row+11, 8),
        (start_row+12, 5, start_row+12, 8),
        (start_row+13, 5, start_row+13, 8),
        (start_row+14, 5, start_row+15, 5),
        (start_row+14, 6, start_row+15, 6),
        (start_row+14, 7, start_row+15, 7),
        (start_row+14, 8, start_row+15, 8),
        (start_row+9, 11, start_row+10, 14),
        (start_row+11, 11, start_row+11, 14),
        (start_row+9, 22, start_row+9, 23),
        (start_row+11, 22, start_row+11, 23),
        (start_row+12, 22, start_row+12, 23),
        (start_row+10, 22, start_row+10, 23),
        (start_row+13, 22, start_row+13, 23),
        (start_row+5,1,start_row+5,3),
        (start_row+5, 4, start_row+5, 11),
        (start_row+5,12,start_row+5,13),
        (start_row+5,14,start_row+5,16),
        (start_row+5,21,start_row+5,24),
        (start_row+6,1,start_row+6,3),
        (start_row+6, 4, start_row+6, 11),
        (start_row+6,12,start_row+6,13),
        (start_row+6,14,start_row+6,16),
        (start_row+6,21,start_row+6,24),
        (start_row+7,1,start_row+7,3),
        (start_row+7, 4, start_row+7, 6),
        (start_row+7, 7, start_row+7, 8),
        (start_row+7, 9, start_row+7, 11),
        (start_row+7,12,start_row+7,13),
        
        (start_row+7,17,start_row+7,18),
        (start_row+7,21,start_row+7,24),
        (start_row+8,1,start_row+8,3),
        (start_row+8, 4, start_row+8, 11),
        (start_row+8,12,start_row+8,13),
        (start_row+8,14,start_row+8,16),
        (start_row+8,17,start_row+8,18),
        (start_row+8,21,start_row+8,24),
    ]
    for merge in merge_cells_instructions:
        ws.merge_cells(start_row=merge[0], start_column=merge[1], end_row=merge[2], end_column=merge[3])

def insert_image(ws, start_row: int):
    img = Image(LOGO_FILE)
    img.width, img.height = 550, 60
    target_cell = f'G{start_row}'
    ws.add_image(img, target_cell)

def add_text_and_styles(ws, start_row: int):
    cells_to_update = [
        ('A4', '地  質  鑽  探  及  土  壤  試  驗  一   覽  表'),
        ('A5', 'SOIL EXPLORATION AND TESTING REPORT'),
        ('A6', '工程名稱：'),
        ('A7', 'Project'),
        ('A8', '鑽孔編號：'),
        ('A9', 'Hole No.'),
        ('L6', '鑽探公司:'),
        ('L7', '座      標：'),
        ('L8', '鑽孔標高：'),
        ('L9', 'Surface Elev.'),
        ('Q8', '地下水位'),
        ('Q9', 'G. W. Depth'),
        ('T6', '地點：'),
        ('T7', 'Location'),
        ('T8', '頁次:'),
        ('T9', 'Page'),
        ('A11', '深度'),
        ('A13', 'Depth'),
        ('A14', '(M)'),
        ('C11', '柱'),
        ('C12', '狀'),
        ('C13', '圖'),
        ('C14', 'Log.'),
        ('D11', '樣號'),
        ('D13', 'Sample'),
        ('D14', 'No.'),
        ('E11', '擊數'),
        ('E12', 'No.of'),
        ('E13', 'Blows'),
        ('E14', 'Per ft.'),
        ('E15', '15cm'),
        ('G8','施工日期：'),
        ('F15', '15cm'),
        ('G15', '15cm'),
        ('H15', 'N值'),
        ('I11', '地質說明'),
        ('I13', 'Soil Description'),
        ('J11', '分類'),
        ('J13', 'USCS'),
        ('J14', 'Classi-'),
        ('J15', 'fication'),
        ('K10', '顆粒分析'),
        ('K12', 'Grain Size Analysis(%)'),
        ('K13', '礫石'),
        ('K16', 'Gravel'),
        ('L13', '砂'),
        ('L16', 'Sand'),
        ('M13', '粉土'),
        ('M16', 'Silt'),
        ('N13', '粘土'),
        ('N16', 'Clay'),
        ('O8','M'),
        ('O10', '自然'),
        ('O11', '含水量'),
        ('O13', 'Water'),
        ('O14', 'Content'),
        ('O16', 'W(%)'),
        ('P10', '比重'),
        ('P13', 'Specific'),
        ('P14', 'Gravity'),
        ('P16', 'G'),
        ('Q10', '當地'),
        ('Q11', '密度'),
        ('Q13', 'Density'),
        ('Q14', 'γt'),
        ('Q16', 'T/M³'),
        ('R10', '空隙比'),
        ('R13', 'Void'),
        ('R14', 'Ratio'),
        ('R16', 'e'),
        ('S10', '塑性'),
        ('S11', '指數'),
        ('S13', 'Liquid'),
        ('S14', 'Limit'),
        ('S16', 'WL(%)'),
        ('T10', '塑性'),
        ('T11', '指數'),
        ('T13', 'Plastic'),
        ('T14', 'Limit'),
        ('T16', 'IP(%)'),
        ('U10', '單軸壓'),
        ('U11', '縮強度'),
        ('U12', 'Uniaxial'),
        ('U13', 'Comp.'),
        ('U14', 'Strength'),
        ('U15', 'qu'),
        ('U16', '(kgf/cm²)'),
        ('V10', '強 度 參 數'),
        ('V12', 'Shear Strength'),
        ('V13', 'Parameter'),
        ('V15', 'ψ'),
        ('V16', '(Degree)'),
        ('W15', "C'"),
        ('W16', '(kgf/cm²)'),
        ('X10', '岩石品'),
        ('X11', '質指標'),
        ('X12', 'Rock'),
        ('X13', 'Quality'),
        ('X14', 'Design-'),
        ('X15', 'ation'),
        ('X16', 'R.Q.D.(%)'),
        # Informatiom of input file-----------------------------------
    ]

    def set_cell_style(cell, align='center'):
        cell.font = Font(name='Times New Roman', size=12, bold=True)
        cell.alignment = Alignment(horizontal=align, vertical='center')

    for cell, value in cells_to_update:
        actual_cell = ws[cell[0] + str(int(cell[1:]) + start_row - 1)]
        actual_cell.value = value
        set_cell_style(actual_cell)

def set_borders(ws, start_row: int):
    for i in range(start_row, start_row + 46, 46):
    # 設置邊框邏輯
        for col in range(1, 25):
            ws.cell(row=i + 8, column=col).border = Border(bottom=MEDIUM_BORDER)
            ws.cell(row=i + 46, column=col).border = Border(top=MEDIUM_BORDER)

        for col in range(2, 24):
            for row in range(i + 17, i + 47):
                ws.cell(row=row, column=col).border = Border(right=THIN_BORDER)

        for row in range(i + 9, i + 46):
            ws.cell(row=row, column=1).border = Border(left=MEDIUM_BORDER)
            ws.cell(row=row, column=24).border = Border(right=MEDIUM_BORDER)

        for row in ws.iter_rows(i + 16, i + 16, 3, 23):
            for cell in row:
                cell.border = Border(top=THIN_BORDER, left=THIN_BORDER)

        for row in ws.iter_rows(i + 15, i + 47, 2, 2):
            for cell in row:
                cell.border = Border(bottom=THIN_BORDER, top=THIN_BORDER, right=THIN_BORDER, left=THIN_BORDER)

        ws[f'A{i + 15}'].border = Border(bottom=THIN_BORDER, left=MEDIUM_BORDER)
        ws[f'X{i + 15}'].border = Border(bottom=THIN_BORDER, right=MEDIUM_BORDER)
        ws[f'X{i + 16}'].border = Border(right=MEDIUM_BORDER,left=THIN_BORDER)

        for row in ws.iter_rows(i + 9, i + 15, 2, 23):
            for cell in row:
                cell.border = Border(right=THIN_BORDER)

        for row in ws.iter_rows(i + 15, i + 15, 3, 23):
            for cell in row:
                cell.border = Border(left=THIN_BORDER, right=THIN_BORDER, bottom=THIN_BORDER)

        for row in ws.iter_rows(i + 10, i + 15, 2, 23):
            for cell in row:
                cell.border = Border(right=THIN_BORDER)

        for row in ws.iter_rows(i + 13, i + 13, 5, 8):
            for cell in row:
                cell.border = Border(bottom=THIN_BORDER, right=THIN_BORDER, left=THIN_BORDER)

        for row in ws.iter_rows(i + 12, i + 12, 11, 14):
            for cell in row:
                cell.border = Border(top=THIN_BORDER, right=THIN_BORDER, left=THIN_BORDER)

        for row in ws.iter_rows(i + 14, i + 14, 22, 23):
            for cell in row:
                cell.border = Border(top=THIN_BORDER, right=THIN_BORDER, left=THIN_BORDER)

def setup_worksheet(ws, start_row: int, project_name: str):
    merge_cells(ws, start_row)
    insert_image(ws, start_row)
    add_text_and_styles(ws, start_row)
    set_borders(ws, start_row)
    adjust_column_width(ws)

def process_worksheet(sheet_name: str, xl: pd.ExcelFile, new_wb: Workbook, project_name: str):
    df = xl.parse(sheet_name)
    Layer = df.iloc[:, 20][4:].dropna().tolist()
    hatch_num = df.iloc[:, 21][4:].dropna().tolist()
    sample_num = df.iloc[:, 1][4:].dropna().tolist()
    sample_depth = df.iloc[:, 0][4:].dropna().tolist() 
    N_value = df.iloc[:, 5][4:].dropna().tolist()
    N1_value = df.iloc[:, 2][4:].tolist()
    N2_value = df.iloc[:, 3][4:].tolist()
    N3_value = df.iloc[:, 4][4:].tolist()
    Classification = df.iloc[:, 13][4:].tolist()
    Gravel = df.iloc[:, 9][4:].tolist() #有問題
    Sand = df.iloc[:, 10][4:].tolist() #有問題
    Silt = df.iloc[:, 11][4:].tolist() #有問題
    Clay = df.iloc[:, 12][4:].tolist() #有問題
    Water_content = df.iloc[:, 14][4:].tolist() #有問題
    Specific_gravity = df.iloc[:, 15][4:].tolist() #有問題
    Density = df.iloc[:, 16][4:].tolist() #有問題
    Void_ratio = df.iloc[:, 17][4:].tolist() #有問題
    Liquid_limit = df.iloc[:, 18][4:].tolist() #有問題
    Plastic_index = df.iloc[:, 19][4:].tolist() #有問題
    ws = new_wb[sheet_name]

    def ensure_list(value):
        if isinstance(value, (int, float)):
            return [value]
        return value if isinstance(value, list) else list(value)

    Layer = ensure_list(Layer)
    hatch_num = ensure_list(hatch_num)
    sample_num = ensure_list(sample_num)
    sample_depth = ensure_list(sample_depth)
    N_value = ensure_list(N_value)
    N1_value = ensure_list(N1_value)
    N2_value = ensure_list(N2_value)
    N3_value = ensure_list(N3_value)
    Classification = ensure_list(Classification)
    Gravel = ensure_list(Gravel)
    Sand = ensure_list(Sand)
    Silt = ensure_list(Silt)
    Clay = ensure_list(Clay)
    Water_content = ensure_list(Water_content)
    Specific_gravity = ensure_list(Specific_gravity)
    Density = ensure_list(Density)
    Void_ratio = ensure_list(Void_ratio)
    Liquid_limit = ensure_list(Liquid_limit)
    Plastic_index = ensure_list(Plastic_index)
    print('garvel',Gravel)

    def format_value(value, decimals=2):
        try:
            return f"{float(value):.{decimals}f}"
        except ValueError:
            return value  # Or handle the case differently if necessary

    # Example usage:
    decimals = 2 # Set your desired number of decimal places here

    Classification = [format_value(x, decimals) for x in Classification]
    Gravel = [format_value(x, 1) for x in Gravel]
    print('Gravel',Gravel)
    Sand = [format_value(x, 1) for x in Sand]
    Silt = [format_value(x, 1) for x in Silt]
    Clay = [format_value(x, 1) for x in Clay]
    Water_content = [format_value(x, 1) for x in Water_content]
    Specific_gravity = [format_value(x, decimals) for x in Specific_gravity]
    Density = [format_value(x, decimals) for x in Density]
    Void_ratio = [format_value(x, decimals) for x in Void_ratio]
    Liquid_limit = [format_value(x, 1) for x in Liquid_limit]
    Plastic_index = [format_value(x, 1) for x in Plastic_index]


    page = math.ceil(max(Layer) / 15)

    for i in range(page):
            setup_worksheet(ws, i * 46 + 1, project_name)

            # Set project name in the specific cell
            project_name_cell = ws[f'D{i * 46 + 6}']
            project_name_cell.value = project_name
            project_name_cell.font = Font(name='Times New Roman', size=12, bold=True)
            project_name_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Set sheet name in the specific cell
            sheet_name_cell = ws[f'D{i * 46 + 8}']
            sheet_name_cell.value = sheet_name
            sheet_name_cell.font = Font(name='Times New Roman', size=12, bold=False)
            sheet_name_cell.alignment = Alignment(horizontal='left', vertical='center')
            

            # Page number

            page_num = str(i + 1)
            page_cell = ws[f'U{i * 46 + 8}']
            page_cell.value = '第'+page_num+'頁'
            page_cell.font = Font(name='Times New Roman', size=12, bold=False)
            page_cell.alignment = Alignment(horizontal='left', vertical='center')
            

            for Layer_depth  in Layer:
                # Layer depth
                Layer_depth = round(Layer_depth, 1)
                if Layer_depth<0.5:
                    Layer_depth = 0.5
                y1=round(Layer_depth/0.5)

                y2=(y1/30)
                if y1%30==0:
                    y2=y2-1
                else:
                    y2=math.floor(y2)

                insert_position = int(y1+(y2+1)*16)
                # Layer_depth_cell = ws[f'A{insert_position}']
                # Layer_depth_cell.value = Layer_depth
                # Layer_depth_cell.font = Font(name='Times New Roman', size=12, bold=False)
                # Layer_depth_cell.alignment = Alignment(horizontal='center', vertical='center')

            for sample_depthes,sample_nums,N,N1,N2,N3,Classi,Gravel,Sand,Silt,Clay,Water_content,Specific_gravity,Density,Void_ratio,Liquid_limit,Plastic_index in zip(sample_depth, sample_num, N_value, N1_value, N2_value, N3_value,Classification,Gravel,Sand,Silt,Clay,Water_content,Specific_gravity,Density,Void_ratio,Liquid_limit,Plastic_index):
                # Sample depth
                # Check contents and types of all variables
                # for var_name in [sample_depth, sample_num, N_value, N1_value, N2_value, N3_value, Classification, Gravel, Sand, Silt, Clay, Water_content, Specific_gravity, Density, Void_ratio, Liquid_limit, Plastic_index]:

                sample_depthes = round(sample_depthes, 1)
                if sample_depthes<0.5:
                    sample_depthes = 0.5
                y1=round(sample_depthes/0.5)

                y2=(y1/30)
                if y1%30==0:
                    y2=y2-1
                else:
                    y2=math.floor(y2)
                insert_position = int(y1+(y2+1)*16)    
                sample_num_cell = ws[f'D{insert_position}']
                sample_num_cell.value = sample_nums
                sample_num_cell.font = Font(name='Times New Roman', size=12, bold=False)
                sample_num_cell.alignment = Alignment(horizontal='center', vertical='center')

                # Sample depth
                sample_depth_cell = ws[f'A{insert_position}']
                sample_depth_cell.value = sample_depthes
                sample_depth_cell.font = Font(name='Times New Roman', size=12, bold=False)
                sample_depth_cell.alignment = Alignment(horizontal='center', vertical='center')


                # N value
                N_cell = ws[f'H{insert_position}']
                N_cell.value = N
                N_cell.font = Font(name='Times New Roman', size=12, bold=False)
                N_cell.alignment = Alignment(horizontal='center', vertical='center')
                    

                # N1 value
                N1_cell = ws[f'E{insert_position}']
                N1_cell.value = N1
                N1_cell.font = Font(name='Times New Roman', size=12, bold=False)
                N1_cell.alignment = Alignment(horizontal='center', vertical='center')

                # N2 value
                N2_cell = ws[f'F{insert_position}']
                N2_cell.value = N2
                N2_cell.font = Font(name='Times New Roman', size=12, bold=False)
                N2_cell.alignment = Alignment(horizontal='center', vertical='center')

                # N3 value
                N3_cell = ws[f'G{insert_position}']
                N3_cell.value = N3
                N3_cell.font = Font(name='Times New Roman', size=12, bold=False)
                N3_cell.alignment = Alignment(horizontal='center', vertical='center')

                # Classification 待修正
                # Classi_str = ', '.join(Classi)
                # Classification_cell = ws[f'L{insert_position}']
                # Classification_cell.value = Classi_str
                # Classification_cell.font = Font(name='Times New Roman', size=12, bold=False)
                # Classification_cell.alignment = Alignment(horizontal='center', vertical='center')
                # print('Classi',Classi)
                for i in enumerate(Classi):
                    Classification_cell = ws[f'J{insert_position}']
                    Classification_cell.value = Classi
                    Classification_cell.font = Font(name='Times New Roman', size=12, bold=False)
                    Classification_cell.alignment = Alignment(horizontal='center', vertical='center')

                # Gravel
                # Gravel_cell = ws[f'M{insert_position}']
                # Gravel_cell.value = Gravel
                # Gravel_cell.font = Font(name='Times New Roman', size=12, bold=False)
                # Gravel_cell.alignment = Alignment(horizontal='center', vertical='center')
                # print(Gravel)
                for i in Gravel:
                    Gravel_cell = ws[f'K{insert_position}']
                    Gravel_cell.value = Gravel
                    Gravel_cell.font = Font(name='Times New Roman', size=12, bold=False)
                    Gravel_cell.alignment = Alignment(horizontal='center', vertical='center')
                    Gravel_cell.number_format = numbers.FORMAT_NUMBER_00
                    print('Gravel', Gravel)

                # Sand
                for i in Sand:
                    Sand_cell = ws[f'L{insert_position}']
                    Sand_cell.value = Sand
                    Sand_cell.font = Font(name='Times New Roman', size=12, bold=False)
                    Sand_cell.alignment = Alignment(horizontal='center', vertical='center')
                    Sand_cell.number_format = numbers.FORMAT_NUMBER_00
                    print('Sand',Sand)


                # Silt
                for i in Silt:
                    Silt_cell = ws[f'M{insert_position}']
                    Silt_cell.value = Silt
                    Silt_cell.font = Font(name='Times New Roman', size=12, bold=False)
                    Silt_cell.alignment = Alignment(horizontal='center', vertical='center')
                    Silt_cell.number_format = numbers.FORMAT_NUMBER_00
                    print('Silt',Silt)
                
                # Clay
                for i in Clay:
                    Clay_cell = ws[f'N{insert_position}']
                    Clay_cell.value = Clay
                    Clay_cell.font = Font(name='Times New Roman', size=12, bold=False)
                    Clay_cell.alignment = Alignment(horizontal='center', vertical='center')
                    Clay_cell.number_format = numbers.FORMAT_NUMBER_00
                    print('Clay',Clay)

                # Water content
                for i in Water_content:
                    Water_content_cell = ws[f'O{insert_position}']
                    Water_content_cell.value = Water_content
                    Water_content_cell.font = Font(name='Times New Roman', size=12, bold=False)
                    Water_content_cell.alignment = Alignment(horizontal='center', vertical='center')
                    Water_content_cell.number_format = numbers.FORMAT_NUMBER_00
                    print('Water_content',Water_content)

                # Specific gravity
                for i in Specific_gravity:
                    Specific_gravity_cell = ws[f'P{insert_position}']
                    Specific_gravity_cell.value = Specific_gravity
                    Specific_gravity_cell.font = Font(name='Times New Roman', size=12, bold=False)
                    Specific_gravity_cell.alignment = Alignment(horizontal='center', vertical='center')
                    Specific_gravity_cell.number_format = numbers.FORMAT_NUMBER_00
                    print('Specific_gravity',Specific_gravity)

                # Density
                for i in Density:
                    Density_cell = ws[f'Q{insert_position}']
                    Density_cell.value = Density
                    Density_cell.font = Font(name='Times New Roman', size=12, bold=False)
                    Density_cell.alignment = Alignment(horizontal='center', vertical='center')
                    Density_cell.number_format = numbers.FORMAT_NUMBER_00
                    print('Density',Density)

                # Void ratio
                for i in Void_ratio:
                    Void_ratio_cell = ws[f'R{insert_position}']
                    Void_ratio_cell.value = Void_ratio
                    Void_ratio_cell.font = Font(name='Times New Roman', size=12, bold=False)
                    Void_ratio_cell.alignment = Alignment(horizontal='center', vertical='center')
                    Void_ratio_cell.number_format = numbers.FORMAT_NUMBER_00
                    print('Void_ratio',Void_ratio)

                # Liquid limit
                for i in Liquid_limit:
                    Liquid_limit_cell = ws[f'S{insert_position}']
                    Liquid_limit_cell.value = Liquid_limit
                    Liquid_limit_cell.font = Font(name='Times New Roman', size=12, bold=False)
                    Liquid_limit_cell.alignment = Alignment(horizontal='center', vertical='center')
                    Liquid_limit_cell.number_format = numbers.FORMAT_NUMBER_00
                    print('Liquid_limit',Liquid_limit)

                # Plastic insex
                for i in Plastic_index:
                    Plastic_index_cell = ws[f'T{insert_position}']
                    Plastic_index_cell.value = Plastic_index
                    Plastic_index_cell.font = Font(name='Times New Roman', size=12, bold=False)
                    Plastic_index_cell.alignment = Alignment(horizontal='center', vertical='center')
                    Plastic_index_cell.number_format = numbers.FORMAT_NUMBER_00
                    print('Plastic_index',Plastic_index)





                



                

                

    

def main():
    app=Application()
    app.mainloop()

if __name__ == "__main__":
    main()