import os
import logging
from typing import List, Tuple
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side
import pandas as pd
import math
import argparse
import tkinter as tk
from tkinter import filedialog
from tkinter import filedialog, messagebox
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import traceback
import win32com.client as win32
import math


# Configuration
INPUT_FILE = ''
OUTPUT_FILE = ''
LOGO_FILE = '萬頂圖樣.png'
MAIN_SHEET = '工作表1'
PROJECT_NAME_CELL = (0,1)

# Constants
THIN_BORDER = Side(border_style='thin')
MEDIUM_BORDER = Side(border_style='medium')

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')



# 配置和常量部分保持不變

# 設置日誌
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel 處理程序")
        self.geometry("300x100")
        self.create_widgets()

    def create_widgets(self):
        self.start_button = tk.Button(self, text="單孔柱狀圖報告建立", command=self.start_processing)
        self.start_button.pack(expand=True)

    def start_processing(self):
        INPUT_FILE, OUTPUT_FILE = self.select_files()
        if not INPUT_FILE or not OUTPUT_FILE:
            messagebox.showinfo("信息", "文件選擇已取消")
            return

        self.process_files(INPUT_FILE, OUTPUT_FILE)

    def select_files(self):
        INPUT_FILE = filedialog.askopenfilename(title="選擇輸入Excel文件", filetypes=[("Excel files", "*.xlsx")])
        if not INPUT_FILE:
            return None, None

        OUTPUT_FILE = filedialog.asksaveasfilename(title="保存輸出Excel文件", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not OUTPUT_FILE:
            return None, None

        return INPUT_FILE, OUTPUT_FILE


    def process_files(self, INPUT_FILE, OUTPUT_FILE):
        try:
            xl, sheet_names, project_name = load_excel_file(INPUT_FILE)
            new_wb = create_new_workbook(sheet_names)

            for sheet_name in sheet_names:
                process_worksheet(sheet_name, xl, new_wb, project_name)

            if os.path.exists(OUTPUT_FILE):
                os.remove(OUTPUT_FILE)
            
            new_wb.save(OUTPUT_FILE)
            messagebox.showinfo("成功", f"成功創建 {OUTPUT_FILE}")
        except Exception as e:
            error_msg = f"發生錯誤:\n{traceback.format_exc()}"
            print(error_msg)  # 輸出到終端
            logging.error(error_msg)  # 記錄到日誌
            messagebox.showerror("錯誤", f"發生錯誤: {str(e)}\n\n詳細信息已輸出到終端")

def parse_arguments():
    parser = argparse.ArgumentParser(description='Process Excel file for geological data.')
    parser.add_argument('--input', default=INPUT_FILE, help='Input Excel file name')
    parser.add_argument('--output', default=OUTPUT_FILE, help='Output Excel file name')
    return parser.parse_args()

def load_excel_file(filename: str) -> Tuple[pd.ExcelFile, List[str], str]:
    try:
        xl = pd.ExcelFile(filename)
        ws = pd.read_excel(xl, MAIN_SHEET, header=None)
        project_name = ws.iloc[PROJECT_NAME_CELL[0], PROJECT_NAME_CELL[1]]  # Access the cell for project name
        print(project_name)
        sheet_names = xl.sheet_names
        sheet_names.remove(MAIN_SHEET)
        return xl, sheet_names, project_name
    except Exception as e:
        logging.error(f"Error loading Excel file: {e}")
        raise


def create_new_workbook(sheet_names: List[str]) -> Workbook:
    new_wb = Workbook()
    new_wb.remove(new_wb.active)
    for sheet_name in sheet_names:
        new_wb.create_sheet(title=sheet_name)
    return new_wb

def adjust_column_width(ws):
    column_widths = {
        'A': 6.11, 'B': 1, 'C': 6, 'D': 8, 'E': 6, 'F': 6, 'G': 6, 'H': 6, 'I': 35,
        'J': 12, 'K': 8, 'L': 8, 'M': 8, 'N': 8, 'O': 9, 'P': 9, 'Q': 9, 'R': 9,
        'S': 9, 'T': 9, 'U': 10, 'V': 9, 'W': 9, 'X': 12
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

def merge_cells(ws, start_row: int):
    merge_cells_instructions = [
        (start_row, 1, start_row+2, 24),
        (start_row+3, 1, start_row+3, 24),
        (start_row+4, 1, start_row+4, 24),
        (start_row+9, 5, start_row+9, 8),
        (start_row+10, 5, start_row+10, 8),
        (start_row+11, 5, start_row+11, 8),
        (start_row+12, 5, start_row+12, 8),
        (start_row+13, 5, start_row+13, 8),
        (start_row+14, 5, start_row+15, 5),
        (start_row+14, 6, start_row+15, 6),
        (start_row+14, 7, start_row+15, 7),
        (start_row+14, 8, start_row+15, 8),
        (start_row+9, 11, start_row+10, 14),
        (start_row+11, 11, start_row+11, 14),
        (start_row+9, 22, start_row+9, 23),
        (start_row+11, 22, start_row+11, 23),
        (start_row+12, 22, start_row+12, 23),
        (start_row+10, 22, start_row+10, 23),
        (start_row+13, 22, start_row+13, 23),
        (start_row+5,1,start_row+5,3),
        (start_row+5, 4, start_row+5, 11),
        (start_row+5,12,start_row+5,13),
        (start_row+5,14,start_row+5,16),
        (start_row+5,21,start_row+5,24),
        (start_row+6,1,start_row+6,3),
        (start_row+6, 4, start_row+6, 11),
        (start_row+6,12,start_row+6,13),
        (start_row+6,14,start_row+6,16),
        (start_row+6,21,start_row+6,24),
        (start_row+7,1,start_row+7,3),
        (start_row+7, 4, start_row+7, 6),
        (start_row+7, 7, start_row+7, 8),
        (start_row+7, 9, start_row+7, 11),
        (start_row+7,12,start_row+7,13),
        (start_row+7,17,start_row+7,18),
        (start_row+7,21,start_row+7,24),
        (start_row+8,1,start_row+8,3),
        (start_row+8, 4, start_row+8, 11),
        (start_row+8,12,start_row+8,13),
        (start_row+8,14,start_row+8,16),
        (start_row+8,17,start_row+8,18),
        (start_row+8,21,start_row+8,24),
    ]
    for merge in merge_cells_instructions:

        ws.merge_cells(start_row=merge[0], start_column=merge[1], end_row=merge[2], end_column=merge[3])

def insert_image(ws, start_row: int):
    img = Image(LOGO_FILE)
    img.width, img.height = 550, 60
    target_cell = f'G{start_row}'
    ws.add_image(img, target_cell)

def add_text_and_styles(ws, start_row: int):
    cells_to_update = [
        ('A4', '地  質  鑽  探  及  土  壤  試  驗  一   覽  表'),
        ('A5', 'SOIL EXPLORATION AND TESTING REPORT'),
        ('A6', '工程名稱：'),
        ('A7', 'Project'),
        ('A8', '鑽孔編號：'),
        ('A9', 'Hole No.'),
        ('L6', '鑽探公司:'),
        ('L7', '座      標：'),
        ('L8', '鑽孔標高：'),
        ('L9', 'Surface Elev.'),
        ('Q8', '地下水位'),
        ('Q9', 'G. W. Depth'),
        ('T6', '地點：'),
        ('T7', 'Location'),
        ('T8', '頁次:'),
        ('T9', 'Page'),
        ('A11', '深度'),
        ('A13', 'Depth'),
        ('A14', '(M)'),
        ('C11', '柱'),
        ('C12', '狀'),
        ('C13', '圖'),
        ('C14', 'Log.'),
        ('D11', '樣號'),
        ('D13', 'Sample'),
        ('D14', 'No.'),
        ('E11', '擊數'),
        ('E12', 'No.of'),
        ('E13', 'Blows'),
        ('E14', 'Per ft.'),
        ('E15', '15cm'),
        ('G8','施工日期：'),
        ('F15', '15cm'),
        ('G15', '15cm'),
        ('H15', 'N值'),
        ('I11', '地質說明'),
        ('I13', 'Soil Description'),
        ('J11', '分類'),
        ('J13', 'USCS'),
        ('J14', 'Classi-'),
        ('J15', 'fication'),
        ('K10', '顆粒分析'),
        ('K12', 'Grain Size Analysis(%)'),
        ('K13', '礫石'),
        ('K16', 'Gravel'),
        ('L13', '砂'),
        ('L16', 'Sand'),
        ('M13', '粉土'),
        ('M16', 'Silt'),
        ('N13', '粘土'),
        ('N16', 'Clay'),
        ('O8','M'),
        ('O10', '自然'),
        ('O11', '含水量'),
        ('O13', 'Water'),
        ('O14', 'Content'),
        ('O16', 'W(%)'),
        ('P10', '比重'),
        ('P13', 'Specific'),
        ('P14', 'Gravity'),
        ('P16', 'G'),
        ('Q10', '當地'),
        ('Q11', '密度'),
        ('Q13', 'Density'),
        ('Q14', 'γt'),
        ('Q16', 'T/M³'),
        ('R10', '空隙比'),
        ('R13', 'Void'),
        ('R14', 'Ratio'),
        ('R16', 'e'),
        ('S10', '塑性'),
        ('S11', '指數'),
        ('S13', 'Liquid'),
        ('S14', 'Limit'),
        ('S16', 'WL(%)'),
        ('T10', '塑性'),
        ('T11', '指數'),
        ('T13', 'Plastic'),
        ('T14', 'Limit'),
        ('T16', 'IP(%)'),
        ('U10', '單軸壓'),
        ('U11', '縮強度'),
        ('U12', 'Uniaxial'),
        ('U13', 'Comp.'),
        ('U14', 'Strength'),
        ('U15', 'qu'),
        ('U16', '(kgf/cm²)'),
        ('V10', '強 度 參 數'),
        ('V12', 'Shear Strength'),
        ('V13', 'Parameter'),
        ('V15', 'ψ'),
        ('V16', '(Degree)'),
        ('W15', "C'"),
        ('W16', '(kgf/cm²)'),
        ('X10', '岩石品'),
        ('X11', '質指標'),
        ('X12', 'Rock'),
        ('X13', 'Quality'),
        ('X14', 'Design-'),
        ('X15', 'ation'),
        ('X16', 'R.Q.D.(%)'),
        # Informatiom of input file-----------------------------------
    ]

    def set_cell_style(cell, align='center'):
        cell.font = Font(name='Times New Roman', size=12, bold=True)
        cell.alignment = Alignment(horizontal=align, vertical='center')

    for cell, value in cells_to_update:
        actual_cell = ws[cell[0] + str(int(cell[1:]) + start_row - 1)]
        actual_cell.value = value
        set_cell_style(actual_cell)


def set_borders(ws, start_row: int):
    # Define border styles
    medium = Side(style='medium')
    thin = Side(style='thin')
    
    medium_border = Border(left=medium, right=medium, top=medium, bottom=medium)
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    left_medium_border = Border(left=medium)
    right_medium_border = Border(right=medium)
    bottom_medium_border = Border(bottom=medium)
    top_medium_border = Border(top=medium)
    right_thin_border = Border(right=thin)
    
    for i in range(start_row, start_row + 46, 46):
        end_row = i + 46
        
        # Apply borders to the entire range
        for row in ws.iter_rows(min_row=i+8, max_row=end_row, min_col=1, max_col=24):
            for cell in row:
                if cell.row == i+8:
                    cell.border = bottom_medium_border
                elif cell.row == end_row:
                    cell.border = top_medium_border
                elif cell.column == 1:
                    cell.border = left_medium_border
                elif cell.column == 24:
                    cell.border = right_medium_border
                elif i+16 <= cell.row < end_row and cell.column > 1:
                    cell.border = right_thin_border
        
        # Special cases
        # ws.merge_cells(start_row=i+16, start_column=3, end_row=i+16, end_column=23)
        # ws[f'C{i+16}'].border = Border(top=thin, left=thin)
        
        for row in range(i+15, end_row):
            ws.cell(row=row, column=2).border = thin_border
        
        ws[f'A{i+15}'].border = Border(bottom=thin, left=medium)
        ws[f'X{i+15}'].border = Border(bottom=thin, right=medium)
        ws[f'X{i+16}'].border = Border(right=medium, left=thin)
        
        for row in range(i+9, i+16):
            for col in range(2, 24):
                ws.cell(row=row, column=col).border = right_thin_border
        
        for col in range(3, 24):
            ws.cell(row=i+15, column=col).border = Border(left=thin, right=thin, bottom=thin)
        
        for col in range(5, 9):
            ws.cell(row=i+13, column=col).border = Border(bottom=thin, right=thin, left=thin)
        
        for col in range(11, 15):
            ws.cell(row=i+12, column=col).border = Border(top=thin, right=thin, left=thin)
        
        for col in range(22, 24):
            ws.cell(row=i+14, column=col).border = Border(top=thin, right=thin, left=thin)



def setup_worksheet(ws, start_row: int, project_name: str):
    merge_cells(ws, start_row)
    # insert_image(ws, start_row)
    add_text_and_styles(ws, start_row)
    set_borders(ws, start_row)
    adjust_column_width(ws)

def process_worksheet(sheet_name: str, xl: pd.ExcelFile, new_wb: Workbook, project_name: str):
    # 讀取Excel數據並轉換為DataFrame
    df = xl.parse(sheet_name)
    
    # 提取各個欄位的數據
    Layer = df.iloc[:, 20][4:].dropna().tolist()
    hatch_num = df.iloc[:, 21][4:].dropna().tolist()
    sample_num = df.iloc[:, 1][4:].dropna().tolist()
    sample_depth = df.iloc[:, 0][4:].dropna().tolist()
    Classi_fication = df.iloc[:, 13][5:].dropna().tolist()
    N_value = df.iloc[:, 5][4:].dropna().tolist()
    N1_value = df.iloc[:, 2][4:].tolist()
    N2_value = df.iloc[:, 3][4:].tolist()
    N3_value = df.iloc[:, 4][4:].tolist()
    
    # 其他需要的數據欄位
    Gravel = df.iloc[:, 9][4:].dropna().tolist()
    Sand = df.iloc[:, 10][4:].dropna().tolist()
    Silt = df.iloc[:, 11][4:].dropna().tolist()
    Clay = df.iloc[:, 12][4:].dropna().tolist()
    Water_content = df.iloc[:, 14][4:].dropna().tolist()
    Gs = df.iloc[:, 15][4:].dropna().tolist()
    Density = df.iloc[:, 16][4:].dropna().tolist()
    Void_ratio = df.iloc[:, 17][4:].dropna().tolist()
    Liquid_limit = df.iloc[:, 18][4:].dropna().tolist()
    Plastic_limit = df.iloc[:, 19][4:].dropna().tolist()

    # 設置新的工作表
    ws = new_wb[sheet_name]

    # 計算頁數
    page = math.ceil(max(Layer) / 15)

    # 開始頁面設置和數據插入
    for i in range(page):
        # 設置頁面基本信息
        setup_worksheet(ws, i * 46 + 1, project_name)
        
        # 設定項目名稱
        project_name_cell = ws[f'D{i * 46 + 6}']
        project_name_cell.value = project_name
        project_name_cell.font = Font(name='Times New Roman', size=12, bold=True)
        project_name_cell.alignment = Alignment(horizontal='left', vertical='center')

        # 設定工作表名稱
        sheet_name_cell = ws[f'D{i * 46 + 8}']
        sheet_name_cell.value = sheet_name
        sheet_name_cell.font = Font(name='Times New Roman', size=12, bold=False)
        sheet_name_cell.alignment = Alignment(horizontal='left', vertical='center')

        # 頁碼
        page_num = str(i + 1)
        page_cell = ws[f'U{i * 46 + 8}']
        page_cell.value = f'第{page_num}頁'
        page_cell.font = Font(name='Times New Roman', size=12, bold=False)
        page_cell.alignment = Alignment(horizontal='left', vertical='center')

        # 確保Layer和hatch_num為列表格式
        if isinstance(Layer, (int, float)):
            Layer = [Layer]
        if isinstance(hatch_num, (int, float)):
            hatch_num = [hatch_num]

        # 插入分層深度和其他數據
        for Layer_depth, hatch in zip(Layer, hatch_num):
            # 處理層深度四捨五入
            Layer_depth = round(Layer_depth, 2)
            if Layer_depth < 0.5:
                Layer_depth = 0.5
            y1 = round(Layer_depth / 0.5)
            y2 = (y1 / 30)
            y2 = math.floor(y2) if y1 % 30 != 0 else y2 - 1
            insert_position = int(y1 + (y2 + 1) * 16)
            
            # 插入層深度到單元格
            Layer_depth_cell = ws[f'A{insert_position}']
            Layer_depth_cell.value = Layer_depth
            Layer_depth_cell.font = Font(name='Times New Roman', size=12)
            Layer_depth_cell.alignment = Alignment(horizontal='center', vertical='center')

            # 根據 hatch 數值生成對應的 WMF 文件名
            file_name = f"hatch_{hatch}.wmf"





            # 處理其他變量並插入資料
            for sample_depthes, sample_nums, N, N1, N2, N3, classi_fiction, G, S, M, C, Wn, gs, density, void_ratio, liquid_limit, plastic_limit in zip(sample_depth, sample_num, N_value, N1_value, N2_value, N3_value, Classi_fication, Gravel, Sand, Silt, Clay, Water_content, Gs, Density, Void_ratio, Liquid_limit, Plastic_limit):
                
                # 處理樣本深度
                sample_depthes = round(sample_depthes, 1)
                if sample_depthes < 0.5:
                    sample_depthes = 0.5
                y1 = round(sample_depthes / 0.5)
                y2 = (y1 / 30)
                y2 = math.floor(y2) if y1 % 30 != 0 else y2 - 1
                insert_position = int(y1 + (y2 + 1) * 16)

                # 插入樣本號碼
                sample_num_cell = ws[f'D{insert_position}']
                sample_num_cell.value = sample_nums
                sample_num_cell.font = Font(name='Times New Roman', size=12, bold=False)
                sample_num_cell.alignment = Alignment(horizontal='center', vertical='center')

                # 插入分類資料
                classi_fiction_cell = ws[f'J{insert_position}']
                classi_fiction_cell.value = classi_fiction
                classi_fiction_cell.font = Font(name='Times New Roman', size=12, bold=False)
                classi_fiction_cell.alignment = Alignment(horizontal='center', vertical='center')

                # 插入 N、N1、N2、N3 值
                ws[f'H{insert_position}'].value = N
                ws[f'H{insert_position}'].font = Font(name='Times New Roman', size=12, bold=False)
                ws[f'H{insert_position}'].alignment = Alignment(horizontal='center', vertical='center')

                ws[f'E{insert_position}'].value = N1
                ws[f'E{insert_position}'].font = Font(name='Times New Roman', size=12, bold=False)
                ws[f'E{insert_position}'].alignment = Alignment(horizontal='center', vertical='center')

                ws[f'F{insert_position}'].value = N2
                ws[f'F{insert_position}'].font = Font(name='Times New Roman', size=12, bold=False)
                ws[f'F{insert_position}'].alignment = Alignment(horizontal='center', vertical='center')

                ws[f'G{insert_position}'].value = N3
                ws[f'G{insert_position}'].font = Font(name='Times New Roman', size=12, bold=False)
                ws[f'G{insert_position}'].alignment = Alignment(horizontal='center', vertical='center')

                # 插入 Gravel, Sand, Silt, Clay, Water_content, Gs, Density, Void_ratio, Liquid_limit, Plastic_limit
                ws[f'K{insert_position}'].value = G
                ws[f'L{insert_position}'].value = S
                ws[f'M{insert_position}'].value = M
                ws[f'N{insert_position}'].value = C
                ws[f'O{insert_position}'].value = Wn
                ws[f'P{insert_position}'].value = gs
                ws[f'Q{insert_position}'].value = density
                ws[f'R{insert_position}'].value = void_ratio
                ws[f'S{insert_position}'].value = liquid_limit
                ws[f'T{insert_position}'].value = plastic_limit

                # 設置單元格格式
                for cell_range in [f'K{insert_position}', f'L{insert_position}', f'M{insert_position}', f'N{insert_position}', f'O{insert_position}', f'P{insert_position}', f'Q{insert_position}', f'R{insert_position}', f'S{insert_position}', f'T{insert_position}']:
                    ws[cell_range].font = Font(name='Times New Roman', size=12, bold=False)
                    ws[cell_range].alignment = Alignment(horizontal='center', vertical='center')

                # 插入圖塊
                
                







def main():
    app=Application()
    app.mainloop()

if __name__ == "__main__":
    main()