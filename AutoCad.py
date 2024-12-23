import pandas as pd
import numpy as np
import xlwings as xw
import tkinter as tk
from tkinter import filedialog
from pyautocad import Autocad, APoint,aDouble
import win32com.client
import pythoncom
from collections import Counter
import openpyxl
import os
import math

file_path = ""
Upper_depth = 5
Lower_depth = 0
distance=25
hole_width=2
Ground_EL=0
num_lists = 0
lists = []
pre_num=0
example_list=[]
example_list_int=[]
all_dict = []
all_distance=[]
all_Ground_EL=[]
#ruler
ruler_top=0
ruler_bottom=0
depest=''

#放大倍數
scale_factor_h=0
scale_factor_w=0

#dictionary

dictionary={'001':'土壤',
            '010':'岩盤',
            '101':'巨礫',
            '102':'粗礫',
            '103':'礫石',
            '104':'砂或礫石質砂(SW 或 SP)',
            '105':'粉土(MH)',
            '106':'高塑性無機性黏土，中至高塑性有'+'\n'+'機黏土(CH 或 OH)',
            '107':'有機黏土',
            '108':'有機粉土',
            '109':'有機砂',
            '110':'泥炭',
            '111':'崩積物;岩屑堆積;碎屑',
            '112':'填方',
            '202':'紅土礫石',
            '206':'黏土質礫石(GC)',
            '207':'粉土質礫石(GM)',
            '222':'礫質砂',
            '223':'粗砂',
            '224':'細砂',
            '225':'凝灰岩',
            '227':'粉土質砂(SM)',
            '228':'黏土質砂(SC)',
            '229':'鈣質砂',
            '242':'砂質粉土(ML)',
            '244':'低塑性黏土質粉土，低塑性有機粉'+'\n'+'土及粉土質黏土(ML 或 OL)',
            '260':'礫質黏土',
            '262':'砂質黏土',
            '264':'低至中塑性無機性黏土，粉土質黏'+'\n'+'土及砂質黏土(CL 或 CL-ML)',
            '266':'紅土',
            '301':'礫岩',
            '302':'角礫岩',
            '303':'砂岩',
            '307':'泥岩',
            '308':'粉砂岩',
            '309':'頁岩',
            '412':'粉砂質砂岩',
            '414':'泥質砂岩',
            '420':'砂岩夾頁岩',
            '424':'砂泥岩互層',
            '426':'砂岩夾頁岩',
            '428':'頁岩夾砂岩',
            '432':'砂質粉砂岩',
            '434':'泥質粉砂岩',
            '442':'砂質泥岩',
            '444':'粉砂質泥岩',
            '446':'砂質頁岩'
            }

# 使用 Tkinter 檔案對話框選擇檔案
root = tk.Tk()
root.title('選擇檔案')
root.geometry('300x200')


def show():
    global file_path,scale_factor_h, scale_factor_w
    file_path = filedialog.askopenfilename()

    root.destroy()  # 選擇檔案後關閉 Tkinter 根視窗

def read_excel_cell(file_path, sheet_name, row_index, col_name):
    # Read Excel cell based on column name
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    col_index = df.iloc[0].tolist().index(col_name)
    value = df.at[row_index, col_index]
    return value

def vtobj(obj):
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_DISPATCH, obj)

def vtfloat(lst):
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, lst)

def add_text(acad, text_value, insert_point, text_height, alignment):
    text = acad.AddText(text_value, insert_point, text_height)
    text.Alignment = alignment
    text.TextAlignmentPoint = insert_point

def add_line(acad, start_point, end_point, line_weight=13):
    line = acad.AddLine(start_point, end_point)
    line.Lineweight = line_weight


btn = tk.Button(root,
                text='開啟檔案',
                font=('Arial', 20, 'bold'),
                command=show
            )
btn.pack()
root.mainloop()

# 自動連線上 CAD
wincad  = win32com.client.Dispatch("AutoCAD.Application")
# wincad  = Autocad(create_if_not_exists=True)

doc = wincad.ActiveDocument
msp = doc.ModelSpace

doc.SetVariable("INSUNITS", 6)

acad =Autocad().ActiveDocument.ModelSpace

# 讀取 Excel 檔案中的所有工作表
xl = pd.ExcelFile(file_path)

# 透過 sheet_names 取得所有工作表的名稱
sheet_names = xl.sheet_names
df = pd.read_excel(file_path)

#比例
keyword='長度比例'
position=[]
# 遍歷 DataFrame 查找關鍵字
for i, row in df.iterrows():
    for j, value in enumerate(row):
        if keyword in str(value):
            position.append((i, j))
        
            # 打印關鍵字的位置
            file_name=os.path.basename(file_path)
            wb = openpyxl.load_workbook(file_path,data_only=True)
            s1=wb.worksheets[0]
            scale_factor_h=(s1.cell(i+2,j+2).value)
        else:
            # 錯誤視窗
            root = tk.Tk()
            root.title("錯誤視窗")
            root.geometry("300x150")
            # 請填入長度比例
            tk.Label(root, text="請填入長度比例",font = ('Times New Roman',16 )).pack(padx=10, pady=15)
            tk.Button(root, text="關閉",font=('Times New Roman',14), command=root.destroy).pack(pady=30)

keyword='寬度比例'
position=[]
# 遍歷 DataFrame 查找關鍵字
for i, row in df.iterrows():
    for j, value in enumerate(row):
        if keyword in str(value):
            position.append((i, j))
        
            # 打印關鍵字的位置
            file_name=os.path.basename(file_path)
            wb = openpyxl.load_workbook(file_path,data_only=True)
            s1=wb.worksheets[0]
            scale_factor_w=(s1.cell(i+2,j+2).value)
        else:
            # 錯誤視窗
            root = tk.Tk()
            root.title("錯誤視窗")
            root.geometry("300x150")
            # 請填入長度比例
            tk.Label(root, text="請填入寬度比例",font = ('Times New Roman',16 )).pack

# 印出每個工作表的資料
p1=(0,0)
p2=(0,0)

y_start=0
y_end=0
y_start_point= APoint(0, y_start)
y_end_point = APoint(0, y_end)




for index, sheet_name in enumerate(sheet_names[1:], start=1):
    #skip the first sheet
    N_1=0
    N_2=0
    E_1=0
    E_2=0
    num_lists+=1    
    print('sheet_name:', sheet_name)
 #------------------------------------------------------------------------------------------------------------------------------------
    #孔頂高
    df = pd.read_excel(xl, sheet_name, header=None)
    Ground_EL= df.iloc[0,5]
    all_Ground_EL.append(Ground_EL)
    # 報錯視窗
    if math.isnan(Ground_EL):
    # 檢查是否已經有一個視窗存在
        if not hasattr(root, 'winfo_exists') or not root.winfo_exists():
            # Create a Tkinter window to display the error message
            root = tk.Tk()
            root.title("錯誤視窗")
            root.geometry("300x150")
            
            # Create and pack the label to display the message
            tk.Label(root, text=f"{sheet_name}請輸入孔位高程", font=('Times New Roman', 16)).pack(padx=10, pady=15)
            
            # Create and pack the button to close the window
            tk.Button(root, text="關閉", font=('Times New Roman', 14), command=root.destroy).pack(pady=30)

            # Run the Tkinter event loop to show the window
            root.mainloop()
#------------------------------------------------------------------------------------------------------------------------------------
    #位置distance
    N_2=df.iloc[1,1]
    E_2 = df.iloc[2, 1]
    
    if pd.isna(N_2) or pd.isna(E_2):
        distance=25
        # 報錯視窗
        root = tk.Tk()
        root.title('報錯視窗')
        # Create and pack the label to display the message
        tk.Label(root, text=f"{sheet_name}沒有經緯度，孔距為25m",font = ('Times New Roman',16 )).pack(padx=10, pady=15)
        # Create and pack the button to close the window
        tk.Button(root, text="關閉",font=('Times New Roman',14), command=root.destroy).pack(pady=30)

    if index==1:
        distance=15
    else:
        distance_=pow(pow(E_2-E_1,2)+pow(N_2-N_1,2),0.5)/1000000*scale_factor_w
        E_1=E_2
        N_1=N_2
        distance=distance+distance_

    all_distance.append(distance)
#------------------------------------------------------------------------------------------------------------------------------------
    # 鑽孔名稱 (hole_width/2*scale_factor))
    text_value = f"{sheet_name}"  # 使用 f-string 來格式化文字
    word_height=2
    #text_width = len(text_value) * 2.5 * scale_factor_w 
    insert_point=APoint((distance+hole_width/2)*scale_factor_w, (Ground_EL+5)*scale_factor_h,(Ground_EL+5)*scale_factor_h) 
    text = acad.AddText(text_value,insert_point, 1*scale_factor_w)
    text.Alignment=13
    text.TextAlignmentPoint = insert_point
    #EL
    text_value='EL  '+str(Ground_EL)
    insert_point=APoint((distance+hole_width/2)*scale_factor_w, (Ground_EL+3)*scale_factor_h,(Ground_EL+3)*scale_factor_h)
    text = acad.AddText(text_value,insert_point, 0.8*scale_factor_w)
    text.Alignment=13
    text.TextAlignmentPoint = insert_point

#------------------------------------------------------------------------------------------------------------------------------------
    #土層深度
    text_value='Depth(m)'
    insert_point=APoint((distance)*scale_factor_w, Ground_EL*scale_factor_h,Ground_EL*scale_factor_h) 
    text=acad.AddText(text_value,insert_point, 0.8*scale_factor_w)
    text.Alignment=14
    text.TextAlignmentPoint = insert_point
#------------------------------------------------------------------------------------------------------------------------------------
    #spt
    text_value='SPT-N'
    #text=acad.AddText(text_value,APoint(index * scale_factor_w*distance+5*scale_factor_w, 2.5*scale_factor_h),2*scale_factor_w)
    insert_point=APoint((distance+hole_width)*scale_factor_w, Ground_EL*scale_factor_h,Ground_EL*scale_factor_h) 
    text=acad.AddText(text_value,insert_point, 0.8*scale_factor_w)
    text.Alignment=12
    text.TextAlignmentPoint = insert_point
    p1 = APoint(scale_factor_w*distance, Ground_EL)
    p2 = APoint(scale_factor_w*distance + 5, Ground_EL)
#------------------------------------------------------------------------------------------------------------------------------------       
    #地下水位
    GWL = df.iloc[1, 5]
    # 如果地下水位是 NaN
    # 報錯視窗
    if math.isnan(GWL):
        # Create a Tkinter window to display the error message
        root = tk.Tk()
        root.title("錯誤視窗")
        root.geometry("300x150")
        
        # Create and pack the label to display the message
        tk.Label(root, text=f"{sheet_name}請輸入地下水位",font = ('Times New Roman',16 )).pack(padx=10, pady=15)
         
        # Create and pack the button to close the window
        tk.Button(root, text="關閉",font=('Times New Roman',14), command=root.destroy).pack(pady=30)

        # Run the Tkinter event loop to show the window
        root.mainloop
    GWL=Ground_EL-GWL
    GWL=round(GWL,2)
    GWL_point=APoint(distance*scale_factor_w-(6*scale_factor_w),GWL*scale_factor_h)
    #水位線
    GWL_point_end=APoint(distance*scale_factor_w-(8*scale_factor_w),GWL*scale_factor_h)
    line=acad.AddLine(GWL_point,GWL_point_end)
    line.LineWeight=5
    #裝飾線
    line=acad.AddLine(APoint(distance*scale_factor_w-(6.5*scale_factor_w),GWL*scale_factor_h-0.2*scale_factor_h),
                    APoint(distance*scale_factor_w-(7.5*scale_factor_w),GWL*scale_factor_h-0.2*scale_factor_h))
    line.LineWeight=5
    line=acad.AddLine(APoint(distance*scale_factor_w-(6.6*scale_factor_w),GWL*scale_factor_h-0.4*scale_factor_h),
                    APoint(distance*scale_factor_w-(7.4*scale_factor_w),GWL*scale_factor_h-0.4*scale_factor_h))
    line.LineWeight=5
    #箭頭
    line=arrow_start=APoint((GWL_point.x+GWL_point_end.x)/2,GWL*scale_factor_h)
    line.LineWeight=5
    line=acad.AddLine(arrow_start,APoint(arrow_start.x+(0.5*scale_factor_h/pow(3,0.5)),arrow_start.y+(0.5*scale_factor_h)))
    line.LineWeight=5
    line=acad.AddLine(arrow_start,APoint(arrow_start.x-(0.5*scale_factor_h/pow(3,0.5)),arrow_start.y+(0.5*scale_factor_h)))
    line.LineWeight=5
    line=acad.AddLine(APoint(arrow_start.x+(0.5*scale_factor_h/pow(3,0.5)),arrow_start.y+(0.5*scale_factor_h)),
                    APoint(arrow_start.x-(0.5*scale_factor_h/pow(3,0.5)),arrow_start.y+(0.5*scale_factor_h)))
    line.LineWeight=5
    next_col_data_str = str(GWL)
    text='G.W.L.'+'  '+next_col_data_str
    insert_point=APoint((GWL_point.x+GWL_point_end.x)/2,arrow_start.y+(0.5*scale_factor_h))
    text=acad.AddText(text,insert_point, 0.6*scale_factor_w)
    text.Alignment=13
    text.TextAlignmentPoint = insert_point
#------------------------------------------------------------------------------------------------------------------------------------ 
    #畫孔位
    #讀取Layer列的數字
    previous_layer = 0
    y1=Ground_EL
    num_element=0
    Layer = df.iloc[:, 20]
    # 如果 Layer 列為空
    # 報錯視窗
    if Layer.isnull().all():
        # Create a Tkinter window to display the error message
        root = tk.Tk()
        root.title("錯誤視窗")
        root.geometry("300x150")
        
        # Create and pack the label to display the message
        tk.Label(root, text=f"{sheet_name}請輸入土層",font = ('Times New Roman',16 )).pack(padx=10, pady=15)
         
        # Create and pack the button to close the window
        tk.Button(root, text="關閉",font=('Times New Roman',14), command=root.destroy).pack(pady=30)

        # Run the Tkinter event loop to show the window
        root.mainloop()
    Layer=Layer[5:]
    Layer=Layer.dropna()

    depth = df.iloc[:, 0]
    # 如果 depth 列為空
    # 報錯視窗
    if depth.isnull().all():
        # Create a Tkinter window to display the error message
        root = tk.Tk()
        root.title("錯誤視窗")
        root.geometry("300x150")
        
        # Create and pack the label to display the message
        tk.Label(root, text=f"{sheet_name}請輸入深度",font = ('Times New Roman',16 )).pack(padx=10, pady=15)
        tk.Button(root, text="關閉",font=('Times New Roman',14), command=root.destroy).pack(pady=30)

    depth=depth[5:]
    depth=depth.dropna()
    hatch_num=df.iloc[:, 21]
    # 如果 hatch_num 列為空
    # 報錯視窗
    if hatch_num.isnull().all():
        # Create a Tkinter window to display the error message
        root = tk.Tk()
        root.title("錯誤視窗")
        root.geometry("300x150")
        
        # Create and pack the label to display the message
        tk.Label(root, text=f"{sheet_name}請輸入土層",font = ('Times New Roman',16 )).pack(padx=10, pady=15)
        tk.Button(root, text="關閉",font=('Times New Roman',14), command=root.destroy).pack(pady=30)

    hatch_num=hatch_num[5:]
    hatch_num=hatch_num.dropna()
    hatch_num=hatch_num.tolist()
    spt_n = df.iloc[:, 5]
    # 如果 spt_n 列為空
    # 報錯視窗
    if spt_n.isnull().all():
        # Create a Tkinter window to display the error message
        root = tk.Tk()
        root.title("錯誤視窗")
        root.geometry("300x150")
        
        # Create and pack the label to display the message
        tk.Label(root, text=f"{sheet_name}請輸入SPT-N",font = ('Times New Roman',16 )).pack(padx=10, pady=15)
        tk.Button(root, text="關閉",font=('Times New Roman',14), command=root.destroy).pack(pady=30)
    spt_n=spt_n[5:]
    data_dict=dict(zip(hatch_num,Layer))
    all_dict.append(data_dict)

    # Layer列數字迭代
    t=0
    # for index, sheet_name in enumerate(sheet_names):
    for layer in Layer:
        print('layer:', layer)
        times=len(Layer)
        y2=Ground_EL-layer
        p1=APoint(distance*scale_factor_w,y1*scale_factor_h)
        p2=APoint((distance+hole_width)*scale_factor_w,y1*scale_factor_h)
        p3=APoint(distance*scale_factor_w,y2*scale_factor_h)
        p4=APoint((distance+hole_width)*scale_factor_w,y2*scale_factor_h)
        y1=y2
        if depest=='' or y2<depest:
            depest=y2
        
        pnts=[p1.x,p1.y,
            p2.x,p2.y,
            p4.x,p4.y,
            p3.x,p3.y,
            p1.x,p1.y]
        #---------------------------------------------------------------------------------------------------------------------
        #填充線
        pnts = vtfloat(pnts)
        sq = msp.AddLightWeightPolyline(pnts)
        sq.Closed = True
        sq.LineWeight=5
        # Convert depth to a numeric type
        #depth = pd.to_numeric(depth, errors='coerce')
        outerLoop = []
        outerLoop.append(sq)
        outerLoop = vtobj(outerLoop)
 
        # 将 hatch_num 转换为整数类型
        
        h=int(hatch_num[t])
        hatchobj = msp.AddHatch(1, h, True)
        hatchobj.PatternScale = 0.5*scale_factor_w 
        hatchobj.AppendOuterLoop(outerLoop)
        hatchobj.Evaluate()
        hatchobj.LineWeight=5
        example_list.append(h)
        
        # Check if depth is not NaN
        #分層深度
        #depth
        Layer_text = f"{layer:.1f}"
        insert_point=APoint((distance-0.5)*scale_factor_w, y2*scale_factor_h,y2*scale_factor_h)
        text = acad.AddText(Layer_text,insert_point, 0.5*scale_factor_w)
        text.Alignment=11  
        text.TextAlignmentPoint = insert_point
        nan_encountered = False

        t+=1
        if t==times:
            break
        #---------------------------------------------------------------------------------------------------------------------
    for depth,spt in zip(depth,spt_n):
        #深度迭代
        if Ground_EL>ruler_top:
            ruler_top=Ground_EL

        if Ground_EL-layer<depest:
            depest=Ground_EL-depth
       
        #spt
        if not pd.isna(spt):
            text_value=spt
            insert_point=APoint((distance+hole_width+0.5)*scale_factor_w, (Ground_EL-depth)*scale_factor_h,(Ground_EL-depth)*scale_factor_h)
            text = acad.AddText(text_value,insert_point, 0.5*scale_factor_w)
            text.Alignment=9
            text.TextAlignmentPoint = insert_point

    pnts_outer=[distance*scale_factor_w,Ground_EL*scale_factor_h,
                (distance+hole_width)*scale_factor_w,Ground_EL*scale_factor_h,
                (distance+hole_width)*scale_factor_w,(Ground_EL-layer)*scale_factor_h,
                distance*scale_factor_w,(Ground_EL-layer)*scale_factor_h,
                distance*scale_factor_w,Ground_EL*scale_factor_h
    ]
    pnts_outer = vtfloat(pnts_outer)
    sq = msp.AddLightWeightPolyline(pnts_outer)
    sq.Closed = True
    sq.LineWeight=5
    outerLoop = []
    outerLoop.append(sq)
    outerLoop = vtobj(outerLoop)
example_list=list(set(example_list))

for i in example_list:
    example_list_int.append(int(i))
count=str(len(example_list_int))

# 將下面這段程式碼加入到你的程式中，用來印出字典中對應的值
# 定義圖例框位置和大小
# 定義圖例框位置和大小
legend_x = -32 * scale_factor_w
legend_y = ruler_top*scale_factor_h
legend_width = 1.5 * scale_factor_w
legend_height = 1.5 * scale_factor_w  # 每個格子的高度

# 加入圖例標題
text = '圖例:'
insert_point = APoint(legend_x, legend_y)
text = acad.AddText(text, insert_point, 1 * scale_factor_w)
text.Alignment = 6  # TopLeft
text.TextAlignmentPoint = insert_point

# 初始化文字和框的Y坐標
text_y = legend_y - 2 * scale_factor_h
box_y = legend_y - 2 * scale_factor_h

# 字和圖例框
for i in example_list_int:
    # 設置文字
    text = dictionary[str(i)]
    text_insert_point = APoint(legend_x + legend_width + 0.5 * scale_factor_w, text_y)
    text_obj = acad.AddMText(text_insert_point, 0.5, text)
    text_obj.Height = 1 * scale_factor_w

    # 設置圖例框
    legend_top_left = APoint(legend_x, box_y)
    legend_top_right = APoint(legend_x + legend_width, box_y)
    legend_bottom_right = APoint(legend_x + legend_width, box_y - legend_height)
    legend_bottom_left = APoint(legend_x, box_y - legend_height)

    pnts = [
        legend_top_left.x, legend_top_left.y,
        legend_top_right.x, legend_top_right.y,
        legend_bottom_right.x, legend_bottom_right.y,
        legend_bottom_left.x, legend_bottom_left.y,
        legend_top_left.x, legend_top_left.y
    ]
    pnts = vtfloat(pnts)

    sq = msp.AddLightWeightPolyline(pnts)
    sq.Closed = True
    sq.LineWeight=5
    outerLoop = [sq]
    outerLoop = vtobj(outerLoop)
    hatchobj = msp.AddHatch(1, i, True)
    hatchobj.PatternScale = 0.5 * scale_factor_w  # 設置填充線比例
    hatchobj.AppendOuterLoop(outerLoop)
    hatchobj.Evaluate()
    hatchobj.LineWeight=5

    # 更新Y坐標
    if i in [106, 244, 264]:
        text_y -= 4 * scale_factor_h
        box_y -= 4 * scale_factor_h
    else:
        text_y -= 2 * scale_factor_h
        box_y -= 2 * scale_factor_h


if not math.isnan(depest):
    ruler_bottom = round(depest - 1)
else:
    ruler_bottom = 0  # or another default value
    # 報錯視窗
    root = tk.Tk()
    root.title('請填入高程Ground EL')

# 深度尺
ruler_top=round(ruler_top+1)
ruler_length = round((ruler_top-ruler_bottom))
insertion_point = APoint(0, ruler_top)
#insert_end=APoint(0,ruler_top-ruler_bottom)
line=acad.AddLine(insertion_point*scale_factor_h, APoint(0,(ruler_top-ruler_length)*scale_factor_h))
line.LineWeight=5
for i in range(ruler_top, ruler_bottom,-1):

    if i % 10 == 0:
        # 画长刻度线
        line=acad.AddLine(APoint(0, i * scale_factor_h), APoint(2 * scale_factor_w, i * scale_factor_h))
        line.LineWeight=5
        text=i/10*10
        insert_point=APoint(3 * scale_factor_w+4, i * scale_factor_h)
        text=acad.AddText(str(text),insert_point,0.7 * scale_factor_w)
        text.Alignment=9
        text.TextAlignmentPoint = insert_point
    elif i % 5 == 0:
        # 画中等长度的刻度线
        line=acad.AddLine(APoint(0, i * scale_factor_h), APoint(1 * scale_factor_w, i * scale_factor_h))
        line.LineWeight=5
        text=i/5*5
        insert_point=APoint(3 * scale_factor_w+4,i * scale_factor_h)
        text=acad.AddText(str(text),insert_point,0.7 * scale_factor_w)
        text.Alignment=9
        text.TextAlignmentPoint = insert_point
    else:
        # 画短刻度线
        line=acad.AddLine(APoint(0, i * scale_factor_h), APoint(0.5 * scale_factor_w, i * scale_factor_h))
        line.LineWeight=5
acad.AddLine(y_start_point,y_end_point)

# If key[i]==key[i+1] then connect the two points

for i in range(len(all_dict) - 1):
    current_dict = all_dict[i]
    next_dict = all_dict[i + 1]
    first_key_current = list(current_dict.keys())[0]
    first_key_next = list(next_dict.keys())[0]
    matched_keys = []
    for key_1 in current_dict:
        for key_2 in next_dict:
            if key_1 == key_2 and key_1 not in matched_keys:
                p1 = APoint((all_distance[i] + hole_width) * scale_factor_w, (all_Ground_EL[i] - float(current_dict[key_1])) * scale_factor_h)
                p2 = APoint((all_distance[i + 1]) * scale_factor_w, (all_Ground_EL[i + 1] - float(next_dict[key_2])) * scale_factor_h)
                line = acad.AddLine(p1, p2)
                line.LineWeight = 5
                matched_keys.append(key_1)

    if first_key_current == first_key_next:
        p1 = APoint((all_distance[i] + hole_width) * scale_factor_w, 
                    (all_Ground_EL[i] * scale_factor_h))
        p2 = APoint((all_distance[i + 1]) * scale_factor_w, 
                    (all_Ground_EL[i + 1] * scale_factor_h))
        line = acad.AddLine(p1, p2)
        line.LineWeight = 5
