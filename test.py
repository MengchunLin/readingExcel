import os
import logging
from typing import List, Tuple
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side
import pandas as pd
import math
import argparse

# Configuration
INPUT_FILE = '七孔測試.xlsx'
OUTPUT_FILE = 'new_single_hole_output.xlsx'
LOGO_FILE = '萬頂圖樣.png'
MAIN_SHEET = '工作表1'
PROJECT_NAME_CELL = (1, 2)

# Constants
THIN_BORDER = Side(border_style='thin')
MEDIUM_BORDER = Side(border_style='medium')

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def parse_arguments():
    parser = argparse.ArgumentParser(description='Process Excel file for geological data.')
    parser.add_argument('--input', default=INPUT_FILE, help='Input Excel file name')
    parser.add_argument('--output', default=OUTPUT_FILE, help='Output Excel file name')
    return parser.parse_args()

def load_excel_file(filename: str) -> Tuple[pd.ExcelFile, List[str], str]:
    try:
        xl = pd.ExcelFile(filename)
        ws = xl[MAIN_SHEET]
        project_name = ws.iloc[PROJECT_NAME_CELL[0], PROJECT_NAME_CELL[1]]
        sheet_names = xl.sheet_names
        sheet_names.remove(MAIN_SHEET)
        return xl, sheet_names, project_name
    except Exception as e:
        logging.error(f"Error loading Excel file: {e}")
        raise

def create_new_workbook(sheet_names: List[str]) -> Workbook:
    new_wb = Workbook()
    new_wb.remove(new_wb.active)
    for sheet_name in sheet_names:
        new_wb.create_sheet(title=sheet_name)
    return new_wb

def adjust_column_width(ws):
    column_widths = {
        'A': 6.11, 'B': 1, 'C': 6, 'D': 8, 'E': 6, 'F': 6, 'G': 6, 'H': 6, 'I': 35,
        'J': 12, 'K': 8, 'L': 8, 'M': 8, 'N': 8, 'O': 9, 'P': 9, 'Q': 9, 'R': 9,
        'S': 9, 'T': 9, 'U': 10, 'V': 9, 'W': 9, 'X': 12
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

def merge_cells(ws, start_row: int):
    merge_cells_instructions = [
        (start_row, 1, start_row+2, 24),
        (start_row+3, 1, start_row+3, 24),
        (start_row+4, 1, start_row+4, 24),
        (start_row+9, 5, start_row+9, 8),
        (start_row+10, 5, start_row+10, 8),
        (start_row+11, 5, start_row+11, 8),
        (start_row+12, 5, start_row+12, 8),
        (start_row+13, 5, start_row+13, 8),
        (start_row+14, 5, start_row+15, 5),
        (start_row+14, 6, start_row+15, 6),
        (start_row+14, 7, start_row+15, 7),
        (start_row+14, 8, start_row+15, 8),
        (start_row+9, 11, start_row+10, 14),
        (start_row+11, 11, start_row+11, 14),
        (start_row+9, 22, start_row+9, 23),
        (start_row+11, 22, start_row+11, 23),
        (start_row+12, 22, start_row+12, 23),
        (start_row+10, 22, start_row+10, 23),
        (start_row+13, 22, start_row+13, 23),
        (start_row+5,1,start_row+5,3),
        (start_row+5, 4, start_row+5, 11),
        (start_row+5,12,start_row+5,13),
        (start_row+5,14,start_row+5,16),
        (start_row+5,21,start_row+5,24),
        (start_row+6,1,start_row+6,3),
        (start_row+6, 4, start_row+6, 11),
        (start_row+6,12,start_row+6,13),
        (start_row+6,14,start_row+6,16),
        (start_row+6,21,start_row+6,24),
        (start_row+7,1,start_row+7,3),
        (start_row+7, 4, start_row+7, 11),
        (start_row+7,12,start_row+7,13),
        (start_row+7,14,start_row+7,16),
        (start_row+7,17,start_row+7,18),
        (start_row+7,21,start_row+7,24),
        (start_row+8,1,start_row+8,3),
        (start_row+8, 4, start_row+8, 11),
        (start_row+8,12,start_row+8,13),
        (start_row+8,14,start_row+8,16),
        (start_row+8,17,start_row+8,18),
        (start_row+8,21,start_row+8,24),
    ]
    for merge in merge_cells_instructions:
        ws.merge_cells(start_row=merge[0], start_column=merge[1], end_row=merge[2], end_column=merge[3])

def insert_image(ws, start_row: int):
    img = Image(LOGO_FILE)
    img.width, img.height = 550, 60
    target_cell = f'G{start_row}'
    ws.add_image(img, target_cell)

def add_text_and_styles(ws, start_row: int, project_name: str):
    cells_to_update = [
        ('A4', '地  質  鑽  探  及  土  壤  試  驗  一   覽  表'),
        ('A5', 'SOIL EXPLORATION AND TESTING REPORT'),
        ('A6', '工程名稱：'),
        ('A7', 'Project'),
        ('A8', '鑽孔編號：'),
        ('A9', 'Hole No.'),
        ('L6', '鑽探公司:'),
        ('L7', '座      標：'),
        ('L8', '鑽孔標高：'),
        ('L9', 'Surface Elev.'),
        ('Q8', '地下水位'),
        ('Q9', 'G. W. Depth'),
        ('T6', '地點：'),
        ('T7', 'Location'),
        ('T8', '頁次:'),
        ('T9', 'Page'),
        ('A11', '深度'),
        ('A13', 'Depth'),
        ('A14', '(M)'),
        ('C11', '柱'),
        ('C12', '狀'),
        ('C13', '圖'),
        ('C14', 'Log.'),
        ('D11', '樣號'),
        ('D13', 'Sample'),
        ('D14', 'No.'),
        ('E11', '擊數'),
        ('E12', 'No.of'),
        ('E13', 'Blows'),
        ('E14', 'Per ft.'),
        ('E15', '15cm'),
        ('F15', '15cm'),
        ('G15', '15cm'),
        ('H15', 'N值'),
        ('I11', '地質說明'),
        ('I13', 'Soil Description'),
        ('J11', '分類'),
        ('J13', 'USCS'),
        ('J14', 'Classi-'),
        ('J15', 'fication'),
        ('K10', '顆粒分析'),
        ('K12', 'Grain Size Analysis(%)'),
        ('K13', '礫石'),
        ('K16', 'Gravel'),
        ('L13', '砂'),
        ('L16', 'Sand'),
        ('M13', '粉土'),
        ('M16', 'Silt'),
        ('N13', '粘土'),
        ('N16', 'Clay'),
        ('O10', '自然'),
        ('O11', '含水量'),
        ('O13', 'Water'),
        ('O14', 'Content'),
        ('O16', 'W(%)'),
        ('P10', '比重'),
        ('P13', 'Specific'),
        ('P14', 'Gravity'),
        ('P16', 'G'),
        ('Q10', '當地'),
        ('Q11', '密度'),
        ('Q13', 'Density'),
        ('Q14', 'γt'),
        ('Q16', 'T/M³'),
        ('R10', '空隙比'),
        ('R13', 'Void'),
        ('R14', 'Ratio'),
        ('R16', 'e'),
        ('S10', '塑性'),
        ('S11', '指數'),
        ('S13', 'Liquid'),
        ('S14', 'Limit'),
        ('S16', 'WL(%)'),
        ('T10', '塑性'),
        ('T11', '指數'),
        ('T13', 'Plastic'),
        ('T14', 'Limit'),
        ('T16', 'IP(%)'),
        ('U10', '單軸壓'),
        ('U11', '縮強度'),
        ('U12', 'Uniaxial'),
        ('U13', 'Comp.'),
        ('U14', 'Strength'),
        ('U15', 'qu'),
        ('U16', '(kgf/cm²)'),
        ('V10', '強 度 參 數'),
        ('V12', 'Shear Strength'),
        ('V13', 'Parameter'),
        ('V15', 'ψ'),
        ('V16', '(Degree)'),
        ('W15', "C'"),
        ('W16', '(kgf/cm²)'),
        ('X10', '岩石品'),
        ('X11', '質指標'),
        ('X12', 'Rock'),
        ('X13', 'Quality'),
        ('X14', 'Design-'),
        ('X15', 'ation'),
        ('X16', 'R.Q.D.(%)'),
        # Informatiom of input file-----------------------------------
        ('D6', project_name),
    ]

    def set_cell_style(cell, align='center'):
        cell.font = Font(name='Times New Roman', size=12, bold=True)
        cell.alignment = Alignment(horizontal=align, vertical='center')

    for cell, value in cells_to_update:
        actual_cell = ws[cell[0] + str(int(cell[1:]) + start_row - 1)]
        actual_cell.value = value
        set_cell_style(actual_cell)

def set_borders(ws, start_row: int):
    for i in range(start_row, start_row + 46, 46):
        # Set top and bottom borders
        for col in range(1, 25):
            ws.cell(row=i + 8, column=col).border = Border(bottom=MEDIUM_BORDER)
            ws.cell(row=i + 46, column=col).border = Border(top=MEDIUM_BORDER)

        # Set right borders
        for col in range(2, 24):
            for row in range(i + 17, i + 47):
                ws.cell(row=row, column=col).border = Border(right=THIN_BORDER)

        # Set left and right borders for the whole block
        for row in range(i + 9, i + 46):
            ws.cell(row=row, column=1).border = Border(left=MEDIUM_BORDER)
            ws.cell(row=row, column=24).border = Border(right=MEDIUM_BORDER)

        # Set bottom borders for specific rows
        for row in ws.iter_rows(i + 16, i + 16, 3, 23):
            for cell in row:
                cell.border = Border(top=THIN_BORDER,
                                     left=THIN_BORDER)

        #ruler
        for row in ws.iter_rows(i + 15, i + 47, 2, 2):
            for cell in row:
                cell.border = Border(bottom=THIN_BORDER, top=THIN_BORDER, right=THIN_BORDER, left=THIN_BORDER)

        # Specific border
        ws[f'A{i + 15}'].border = Border(bottom=THIN_BORDER, left=MEDIUM_BORDER)
        ws[f'X{i + 15}'].border = Border(bottom=THIN_BORDER, right=MEDIUM_BORDER)
        ws[f'X{i + 16}'].border = Border(right=MEDIUM_BORDER,left=THIN_BORDER)

        for row in ws.iter_rows(i + 9, i + 15, 2, 23):
            for cell in row:
                cell.border = Border(right=THIN_BORDER)

        # Adjust columns and rows
        for row in ws.iter_rows(i + 15, i + 15, 3, 23):
            for cell in row:
                cell.border = Border(left=THIN_BORDER, right=THIN_BORDER, bottom=THIN_BORDER)

        for row in ws.iter_rows(i + 10, i + 15, 2, 23):
            for cell in row:
                cell.border = Border(right=THIN_BORDER)
                
        # Single area border
        for row in ws.iter_rows(i + 13, i + 13, 5, 8):
            for cell in row:
                cell.border = Border(bottom=THIN_BORDER, right=THIN_BORDER, left=THIN_BORDER)

        for row in ws.iter_rows(i + 12, i + 12, 11, 14):
            for cell in row:
                cell.border = Border(top=THIN_BORDER, right=THIN_BORDER, left=THIN_BORDER)

        for row in ws.iter_rows(i + 14, i + 14, 22, 23):
            for cell in row:
                cell.border = Border(top=THIN_BORDER, right=THIN_BORDER, left=THIN_BORDER)

def setup_worksheet(ws, start_row: int, project_name: str):
    merge_cells(ws, start_row)
    insert_image(ws, start_row)
    add_text_and_styles(ws, start_row, project_name)
    set_borders(ws, start_row)
    adjust_column_width(ws)

def process_worksheet(sheet_name: str, xl: pd.ExcelFile, new_wb: Workbook, project_name: str):
    df = xl.parse(sheet_name)
    Layer = df.iloc[:, 20][5:].dropna().tolist()
    hatch_num = df.iloc[:, 21][5:].dropna().tolist()

    ws = new_wb[sheet_name]
    page = math.ceil(max(Layer) / 14.5)
    for i in range(page):
        setup_worksheet(ws, i * 46 + 1, project_name)

def main():
    args = parse_arguments()
    
    try:
        xl, sheet_names, project_name = load_excel_file(args.input)
        new_wb = create_new_workbook(sheet_names)

        for sheet_name in sheet_names:
            process_worksheet(sheet_name, xl, new_wb, project_name)

        if os.path.exists(args.output):
            os.remove(args.output)
        
        new_wb.save(args.output)
        logging.info(f"Successfully created {args.output}")
    except Exception as e:
        logging.error(f"An error occurred: {e}")

if __name__ == "__main__":
    main()